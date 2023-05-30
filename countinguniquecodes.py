from openpyxl import Workbook, load_workbook
import json
from findSchool import findSchool
from joiningjson import exporting
import multiprocessing
from math import ceil, floor
import json
import string
import random

data = {}

def divided(ws, s, e, board, code, pid):
    data = {}
    ans = 0
    try:
        for row in range(s, e+1):
            if (ws.cell(row,board).value == 'Central Board of Secondary Education'):
                percentage = (row - s) / (e - s) * 100
                outputs = f"ID: {pid} - {floor(percentage)}%"
                print(outputs)
                ans += 1
                if (data.get(ws.cell(row, code).value)):
                    continue
                else:
                    data[ws.cell(row,code).value] = findSchool(ws.cell(row,code).value)
        s = ''.join(random.choices(string.ascii_lowercase + string.ascii_uppercase + string.digits, k = 10))
        with open('./jsonfiles/temporary/' + s + '.json','w') as f:
            json.dump(data, f, indent=4)
        print(ans)
    except:
        s = ''.join(random.choices(string.ascii_lowercase + string.ascii_uppercase + string.digits, k=10))
        with open('./jsonfiles/temporary/' + s + '.json','w') as f:
            json.dump(data, f, indent=4)
        print(ans)

try:
    wb = load_workbook('2020-school-data.xlsx')
    ws = wb.active
    columnNames = []

    for cell in ws[1]:
        columnNames.append(cell.value)
    n = len(columnNames)

    board = columnNames.index('Board/University') + 1
    code = columnNames.index('School Code') + 1
    # print(columnNames[board-1])
    # print(columnNames[code-1])

    ws.cell(1,n+1).value = 'Names'

    rowCount = 5000
    start = 2
    cores = 8
    lenForEachCore = ceil((rowCount-start)/cores)
    processes = []
    pid = 0
    # while (rowCount>start - 1):
    #     s = 0
    #     e = 0
    #     if (rowCount - lenForEachCore <= start -1):
    #         s = start
    #     else:
    #         s = rowCount - lenForEachCore
    #     e = rowCount
    #     rowCount = s - 1
    #     p = multiprocessing.Process(target=divided, args=(ws,s,e, board, code, pid))
    #     p.start()
    #     processes.append(p)
    #     pid += 1
    # for process in processes:
    #     process.join()

    
    # output = exporting()
    # print(f'File Generated : {output}')
    # for row in range(2, 20):
    #     if (ws.cell(row,board).value == 'Central Board of Secondary Education'):
    #         if (data.get(ws.cell(row, code).value)):
    #             ws.cell(row,n+1).value = data[ws.cell(row,code).value]
    #         else:
    #             data[ws.cell(row,code).value] = findSchool(ws.cell(row,code).value)
    #             ws.cell(row,n+1).value = data[ws.cell(row,code).value]
    #     else:
    #         ws.cell(row,n+1).value = 'none'




    # wb.save('temp.xlsx')
    # with open('asdf.json', 'w') as j:
    #     json.dump(data, j, indent=4)
    data = {}
    ans = 0
    s = 2
    e = 5000
    ulis = []
    for row in range(s, e+1):
        if (ws.cell(row,board).value == 'Central Board of Secondary Education'):
            ulis.append(ws.cell(row,code).value)
            # if (data.get(ws.cell(row, code).value)):
            #     continue
            # else:
            #     data[ws.cell(row,code).value] = findSchool(ws.cell(row,code).value)
    uset = set(ulis)
    print(len(uset))
    
except:
    # wb.save('temp.xlsx')
    with open('asdf.json', 'w') as j:
        json.dump(data, j, indent=4)