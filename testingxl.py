from openpyxl import Workbook, load_workbook
import json
from findSchool import findSchool

data = {}

try:
    wb = load_workbook('2020-school-data.xlsx')
    ws = wb.active
    columnNames = []

    for cell in ws[1]:
        columnNames.append(cell.value)
    n = len(columnNames)

    board = columnNames.index('Board/University') + 1
    code = columnNames.index('School Code') + 1
    # print(columnNames[board-1])
    # print(columnNames[code-1])

    ws.cell(1,n+1).value = 'Names'

    for row in range(2, 20):
        if (ws.cell(row,board).value == 'Central Board of Secondary Education'):
            if (data.get(ws.cell(row, code).value)):
                ws.cell(row,n+1).value = data[ws.cell(row,code).value]
            else:
                data[ws.cell(row,code).value] = findSchool(ws.cell(row,code).value)
                ws.cell(row,n+1).value = data[ws.cell(row,code).value]
        else:
            ws.cell(row,n+1).value = 'none'




    wb.save('temp.xlsx')
    with open('asdf.json', 'w') as j:
        json.dump(data, j, indent=4)
except:
    wb.save('temp.xlsx')
    with open('asdf.json', 'w') as j:
        json.dump(data, j, indent=4)